"""Script voor Waterschap Noorderzijlvest voor het bewerken van lab-formulieren."""

import environment
environment.activate()
from openpyxl import load_workbook, writer, worksheet
import click
from pathlib import Path
from copy import copy
import string

import warnings
warnings.filterwarnings("ignore")

"""
To Do:
    - Merged cells:
        merged_cells = copy(ws.merged_cells)
        weggooien for rng in merged_cells.ranges:
            ws.unmerge_cells(rng.coord)
        dan weer terugzetten
    - datavalidatie:
         tree = ws.data_validations.to_tree()
         de range aanpassen in de tree
         de datavalidations opnieuw bouwen met openpyxl.worksheet.datavalidation.DataValidationList.from_tree(tree)
    - conditional formatting
        ook via trees
"""


def insert_row(ws: worksheet, line_number: int) -> worksheet:
    """Insert a row at Excel line number."""
    sheet_lower_right = ws.dimensions.split(":")[1]
    move_range = f"A{line_number}:{sheet_lower_right}"
    ws.move_range(move_range, rows=1, translate=True)
    for idx, cell in enumerate(ws[line_number]):
        cell.fill = copy(ws[line_number+1][idx].fill)
        cell.alignment = copy(ws[line_number+1][idx].alignment)
        cell.font = copy(ws[line_number+1][idx].font)
        cell.border = copy(ws[line_number+1][idx].border)
    return ws


def delete_row(ws: worksheet, line_number: int) -> worksheet:
    """Delete a row at Exel line number."""
    sheet_lower_right = ws.dimensions.split(":")[1]
    move_range = f"A{line_number}:{sheet_lower_right}"
    ws.move_range(move_range, rows=-1, translate=True)
    return ws


def replace(ws: worksheet, replace_list: list) -> worksheet:
    """Search a row and delete or replace dending on new_row."""
    old_value = replace_list[0].split(",")
    new_value = replace_list[1].split(",")

    if len(old_value) == 1:
        old_value = old_value[0]
        new_value = new_value[0]
        for r in range(1, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                s = ws.cell(r, c).value
                if s == old_value:
                    ws.cell(r, c).value = s.replace(old_value, new_value)
                    print((f'"{old_value}" vervangen voor "{new_value}" in rij {r},'
                           f'kolom {c}'))
    elif len(old_value) > 1:
        for r in ws.iter_rows():
            vals = [r[idx].value for idx in range(len(old_value))]
            if vals == old_value:
                for idx, val in enumerate(new_value):
                    r[idx].value = val

    return ws


def multiply(ws: worksheet, ident: str, value: float) -> worksheet:
    """Search a row and delete or replace dending on new_row."""
    for r in ws.iter_rows():
        if r[0].value == ident:
            if r[1].value is not None:
                new_value = r[1].value * value
                print(f'"{ident}": {r[1].value} vervangen voor "{ident}": {new_value}')
                r[1].value = new_value

    return ws


@click.command()
@click.option('--vervang',
              type=str,
              help=('Een vervanginging gescheiden door een ","'
                    '\n\nVoorbeelden:\n\n--vervang "Doorzicht in cm,Doorzicht in dm"'
                    '\n\n--vervang "[meetdoel,KRW],[meetdoel,projectmatig]"'))
@click.option('--vermenigvuldigen',
              type=str,
              help=('Een variabele en een vermenigvuldiging gescheiden met een "," '
                    'tussen de variabele en de vermendigvulding  ""\n\nVoorbeeld:\n\n' 
                    '--vermenigvuldigen "Doorzicht in cm,0.1"'))

def main(vervang, vermenigvuldigen):
    """Invoeren van regels in Excel."""
    input_dir = Path("input").absolute()
    output_dir = Path("output").absolute()
    output_dir.mkdir(exist_ok=True)
    for excel_file in input_dir.glob("*.xlsx"):
        file_name = excel_file.name
        print(file_name)
        out_file = output_dir.joinpath(file_name)
        if out_file.exists():
            out_file.unlink()
        book = load_workbook(excel_file)
        ws = book.worksheets[0]
        if vermenigvuldigen:
            manipulate_list = vermenigvuldigen.split(",")
            ident = manipulate_list[0]
            value = float(manipulate_list[1])
            ws = multiply(ws, ident, value)
        if vervang:
            replace_list = vervang.split(",")
            ws = replace(ws, replace_list)
        writer.excel.save_workbook(book, out_file)


if __name__ == '__main__':
    main()
